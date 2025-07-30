from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook
from datetime import datetime
import tkinter as tk

def coletador_dados_tempo():
    try:
        # Configuração do Chrome para evitar bloqueios
        options = webdriver.ChromeOptions()
        options.add_argument("--disable-blink-features=AutomationControlled")  # Remove a detecção de automação
        options.add_experimental_option("excludeSwitches", ["enable-automation"])  # Oculta aviso de automação
        options.add_experimental_option("useAutomationExtension", False)  # Evita extensões indesejadas

        # Iniciar o navegador com as opções configuradas
        navegador = webdriver.Chrome(options=options)

        # Acessa o Google e busca a previsão do tempo
        navegador.get("https://www.google.com")
        navegador.find_element(By.NAME, 'q').send_keys('Temperatura são paulo', Keys.ENTER)
        navegador.implicitly_wait(5)

        # Coleta os dados de temperatura e umidade
        temperatura = navegador.find_element(By.XPATH, '//*[@id="wob_tm"]').text
        umidade = navegador.find_element(By.XPATH, '//*[@id="wob_hm"]').text
        navegador.quit()

        # Exibe os dados no terminal
        print(f"A temperatura atual de São Paulo é {temperatura}°C e a umidade é {umidade}")

        # Nome do arquivo Excel
        arquivo = 'ProjetoTempo.xlsx'

        # Verifica se o arquivo existe, senão cria um novo
        try:
            wb = load_workbook(arquivo)
            sheet = wb.active
        except FileNotFoundError:
            wb = Workbook()
            sheet = wb.active
            sheet.append(["Temperatura (°C)", "Umidade (%)", "Hora", "Data"])

        # Obtém data e hora atual
        data_atual = datetime.now().strftime('%d-%m-%Y')
        hora_atual = datetime.now().strftime('%H:%M:%S')

        # Adiciona os dados à planilha
        sheet.append([temperatura, umidade, hora_atual, data_atual])
        wb.save(arquivo)

        print("Dados armazenados na planilha com sucesso.")

    except Exception as e:
        print(f"Erro ao buscar os dados: {e}")

# Interface gráfica com Tkinter
class Aplicacao:
    def __init__(self):
        self.layout = tk.Tk()
        self.layout.title("Previsão do Tempo - São Paulo")
        self.layout.geometry("350x100")

        self.descricao = tk.Label(self.layout, text="Atualizar temperatura na planilha:", font=("Arial", 12))
        self.descricao.pack(pady=5)

        self.salvar = tk.Button(self.layout, text="Buscar previsão", command=coletador_dados_tempo, font=("Arial", 10))
        self.salvar.pack(pady=2)

        self.layout.mainloop()

executar_app = Aplicacao() 
